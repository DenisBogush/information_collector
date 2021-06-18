# -*- coding: utf-8 -*-


from openpyxl.utils import get_column_letter

RUB_CURRENCY = '_-* #,##0.00\ "₽"_-;\-* #,##0.00\ "₽"_-;_-* "-"??\ "₽"_-;_-@_-'


def get_col_by_name(sheet, col_name):
    return next(filter(lambda col: col[0].value == col_name, sheet.columns))


# устанавливает формат данных столбцам
def set_format_to_column(sheet, *col_names, fmt=RUB_CURRENCY):
    for c_name in col_names:
        for cell in get_col_by_name(sheet, c_name):
            cell.number_format = fmt


# функция устанавливает автосайз столбцам
def set_autosize(sheet):
    column_letters = tuple(get_column_letter(col_number + 1) for col_number in range(sheet.max_column))
    for column_letter in column_letters:
        sheet.column_dimensions[column_letter].auto_size = True


# функция ищет отношение между  column_dividend и column_divisor и помещает результат в column_result
def share_exchange_rates(column_dividend, column_divisor, column_result):
    for row_number in range(1, len(column_dividend)):
        if column_dividend[row_number].value and column_divisor[row_number].value:
            column_result[row_number].value = round(
                column_dividend[row_number].value / column_divisor[row_number].value, 4)
